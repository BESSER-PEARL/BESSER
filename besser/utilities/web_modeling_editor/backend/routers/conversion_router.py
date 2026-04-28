"""
Conversion Router

Handles all conversion, export, and import endpoints for the BESSER web modeling editor backend.
"""

import ast
import asyncio
import logging
import os
import tempfile
import uuid
import importlib.util
import json
from copy import deepcopy
from datetime import datetime, timezone

from fastapi import APIRouter, HTTPException, File, UploadFile, Body, Form
from fastapi.responses import Response

# Backend response models
from besser.utilities.web_modeling_editor.backend.models.responses import (
    DiagramExportResponse,
    ProjectExportResponse,
)

# BESSER image-to-UML and BUML utilities
from besser.utilities.image_to_buml import image_to_buml
from besser.utilities.kg_to_buml import kg_to_buml
from besser.utilities.owl_to_buml import owl_file_to_knowledge_graph

# BESSER KG → B-UML (deterministic, offline)
from besser.BUML.notations.kg_to_buml import (
    kg_to_class_diagram as kg_to_class_diagram_buml,
    kg_to_object_diagram as kg_to_object_diagram_buml,
)

# BESSER utilities
from besser.utilities.buml_code_builder.domain_model_builder import domain_model_to_code
from besser.utilities.buml_code_builder.agent_model_builder import agent_model_to_code
from besser.utilities.buml_code_builder.project_builder import project_to_code
from besser.utilities.buml_code_builder.state_machine_builder import state_machine_to_code

# Backend models
from besser.utilities.web_modeling_editor.backend.models import (
    DiagramInput,
    ProjectInput,
)

# Backend services - Converters
from besser.utilities.web_modeling_editor.backend.services.converters import (
    # JSON to BUML converters
    process_class_diagram,
    process_state_machine,
    process_agent_diagram,
    process_object_diagram,
    process_kg_diagram,
    json_to_buml_project,
    # BUML to JSON converters
    class_buml_to_json,
    parse_buml_content,
    state_machine_to_json,
    agent_buml_to_json,
    gui_buml_to_json,
    object_model_to_json,
    project_to_json,
    kg_to_json,
)

# Backend services - Other services
from besser.utilities.web_modeling_editor.backend.services.utils.agent_generation_utils import (
    extract_openai_api_key,
)
from besser.utilities.web_modeling_editor.backend.services.reverse_engineering import (
    csv_to_domain_model,
)

# Backend configuration
from besser.utilities.web_modeling_editor.backend.config import (
    get_generator_info,
)

# Backend constants
from besser.utilities.web_modeling_editor.backend.constants.constants import (
    API_VERSION,
    TEMP_DIR_PREFIX,
    AGENT_TEMP_DIR_PREFIX,
    CSV_TEMP_DIR_PREFIX,
    OUTPUT_DIR_NAME,
    AGENT_MODEL_FILENAME,
)

# Centralized error handling
from besser.utilities.web_modeling_editor.backend.routers.error_handler import (
    handle_endpoint_errors,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# File upload size limits
# ---------------------------------------------------------------------------
MAX_CSV_SIZE = 5 * 1024 * 1024       # 5 MB
MAX_IMAGE_SIZE = 10 * 1024 * 1024     # 10 MB
MAX_BUML_SIZE = 2 * 1024 * 1024       # 2 MB
MAX_OWL_SIZE = 5 * 1024 * 1024        # 5 MB

# ---------------------------------------------------------------------------
# Allowed MIME / extension sets per upload type
# ---------------------------------------------------------------------------
ALLOWED_SPREADSHEET_EXTENSIONS = {".csv", ".xlsx"}
ALLOWED_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg"}
ALLOWED_BUML_EXTENSIONS = {".py"}
ALLOWED_KG_EXTENSIONS = {".ttl", ".rdf", ".json"}
ALLOWED_OWL_EXTENSIONS = {".owl", ".ttl", ".rdf", ".xml", ".nt", ".n3"}


def _validate_upload(file: UploadFile, *, max_size: int, allowed_extensions: set[str], content: bytes) -> None:
    """Validate an uploaded file's size and extension.

    Raises ``HTTPException`` with status 413 (payload too large) or 415
    (unsupported media type) when the check fails.
    """
    # --- size check ---
    if len(content) > max_size:
        max_mb = max_size / (1024 * 1024)
        raise HTTPException(
            status_code=413,
            detail=f"File too large. Maximum allowed size is {max_mb:.0f} MB.",
        )

    # --- extension check ---
    filename = file.filename or ""
    _, ext = os.path.splitext(filename)
    if ext.lower() not in allowed_extensions:
        raise HTTPException(
            status_code=415,
            detail=(
                f"Unsupported file type '{ext}'. "
                f"Allowed extensions: {', '.join(sorted(allowed_extensions))}."
            ),
        )


def _validate_file_content(content: bytes, filename: str) -> None:
    """Validate uploaded file content based on its extension.

    Raises ``HTTPException`` with status 400 when the content does not match
    the expected format for the given file type.
    """
    _, ext = os.path.splitext(filename)
    ext = ext.lower()

    if ext == ".csv":
        # Must be valid UTF-8 text with at least one line containing comma or semicolon separators
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            raise HTTPException(status_code=400, detail="CSV file is not valid UTF-8 text.")
        lines = [line for line in text.splitlines() if line.strip()]
        if not lines:
            raise HTTPException(status_code=400, detail="CSV file is empty.")
        has_separator = any("," in line or ";" in line for line in lines)
        if not has_separator:
            raise HTTPException(
                status_code=400,
                detail="CSV file does not contain comma or semicolon separators.",
            )

    elif ext == ".xlsx":
        # XLSX files are ZIP archives — check the ZIP magic bytes.
        if not content.startswith(b"PK\x03\x04"):
            raise HTTPException(
                status_code=400,
                detail="XLSX file appears to be corrupted or is not a valid Excel workbook.",
            )

    elif ext == ".py":
        # Must be valid Python syntax
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            raise HTTPException(status_code=400, detail="Python file is not valid UTF-8 text.")
        try:
            ast.parse(text, filename=filename)
        except SyntaxError as e:
            raise HTTPException(
                status_code=400,
                detail=f"Python file contains invalid syntax: {e.msg} (line {e.lineno}).",
            )

    elif ext in (".png", ".jpg", ".jpeg"):
        # Check magic bytes
        if ext == ".png":
            if not content.startswith(b'\x89PNG'):
                raise HTTPException(
                    status_code=400,
                    detail="File content does not match PNG format (invalid magic bytes).",
                )
        else:  # .jpg / .jpeg
            if not content.startswith(b'\xff\xd8\xff'):
                raise HTTPException(
                    status_code=400,
                    detail="File content does not match JPEG format (invalid magic bytes).",
                )

    elif ext in (".owl", ".ttl", ".rdf", ".xml", ".nt", ".n3"):
        # RDF / OWL: must be non-empty and decodable as UTF-8. The real parser
        # (rdflib) performs the structural validation.
        if not content.strip():
            raise HTTPException(status_code=400, detail="OWL/RDF file is empty.")
        try:
            content.decode("utf-8")
        except UnicodeDecodeError:
            raise HTTPException(status_code=400, detail="OWL/RDF file is not valid UTF-8 text.")


async def _read_file(path: str, mode: str = "r", **kwargs) -> str | bytes:
    """Read a file without blocking the event loop."""
    def _do_read():
        with open(path, mode, **kwargs) as fh:
            return fh.read()
    return await asyncio.to_thread(_do_read)


async def _write_file(path: str, data: bytes | str, mode: str = "wb") -> None:
    """Write data to a file without blocking the event loop."""
    def _do_write():
        with open(path, mode) as fh:
            fh.write(data)
    await asyncio.to_thread(_do_write)


def _xlsx_bytes_to_csv_file(xlsx_bytes: bytes, csv_path: str) -> None:
    """Convert the first worksheet of an XLSX workbook to a UTF-8 CSV file.

    Lazy-imports openpyxl so the rest of the backend stays importable if the
    optional dependency is missing.
    """
    import csv as _csv
    import io

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="XLSX support requires the 'openpyxl' package. Install it with `pip install openpyxl`.",
        ) from exc

    try:
        workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail=f"Could not read XLSX file: {exc}",
        ) from exc

    try:
        worksheet = workbook.active
        if worksheet is None:
            raise HTTPException(status_code=400, detail="XLSX file contains no worksheets.")

        with open(csv_path, "w", encoding="utf-8", newline="") as fh:
            writer = _csv.writer(fh)
            rows_written = 0
            for row in worksheet.iter_rows(values_only=True):
                writer.writerow(["" if cell is None else cell for cell in row])
                rows_written += 1
    finally:
        workbook.close()

    if rows_written == 0:
        raise HTTPException(status_code=400, detail="XLSX worksheet is empty.")


async def _read_json_file(path: str) -> dict:
    """Read and parse a JSON file without blocking the event loop."""
    def _do_read():
        with open(path, "r", encoding="utf-8") as fh:
            return json.load(fh)
    return await asyncio.to_thread(_do_read)


router = APIRouter(prefix="/besser_api", tags=["conversion"])


@router.post("/export-project-as-buml")
@handle_endpoint_errors("export_project_as_buml")
async def export_project_as_buml(input_data: ProjectInput = Body(...)):
    buml_project = json_to_buml_project(input_data)

    with tempfile.TemporaryDirectory(prefix=f"{TEMP_DIR_PREFIX}{uuid.uuid4().hex}_") as temp_dir:
        output_file_path = os.path.join(temp_dir, "project.py")
        project_to_code(project=buml_project, file_path=output_file_path)
        file_content = await _read_file(output_file_path, "r", encoding="utf-8")
        return Response(
            content=file_content,
            media_type="text/plain",
            headers={"Content-Disposition": 'attachment; filename="project.py"'},
        )


@router.post("/export-buml")
@handle_endpoint_errors("export_buml")
async def export_buml(input_data: DiagramInput):
    # Create unique temporary directory for this request
    with tempfile.TemporaryDirectory(prefix=f"{TEMP_DIR_PREFIX}{uuid.uuid4().hex}_") as temp_dir:
        json_data = input_data.model_dump()
        elements_data = input_data.model
        if elements_data.get("type") == "StateMachineDiagram":
            state_machine = process_state_machine(json_data)
            output_file_path = os.path.join(temp_dir, "state_machine.py")
            state_machine_to_code(model=state_machine, file_path=output_file_path)
            file_content = await _read_file(output_file_path, "rb")
            return Response(
                content=file_content,
                media_type="text/plain",
                headers={
                    "Content-Disposition": 'attachment; filename="state_machine.py"'
                },
            )

        elif elements_data.get("type") == "ClassDiagram":
            buml_model = process_class_diagram(json_data)
            output_file_path = os.path.join(temp_dir, "domain_model.py")
            domain_model_to_code(model=buml_model, file_path=output_file_path)
            file_content = await _read_file(output_file_path, "rb")
            return Response(
                content=file_content,
                media_type="text/plain",
                headers={"Content-Disposition": 'attachment; filename="domain_model.py"'},
            )

        elif elements_data.get("type") == "ObjectDiagram":
            # Build the reference class diagram model from referenceDiagramData
            reference_data = input_data.referenceDiagramData or elements_data.get("referenceDiagramData", {})
            if not reference_data:
                raise ValueError("Object diagram requires reference class diagram data")

            reference_json = {"title": "Reference Classes", "model": reference_data}
            buml_model = process_class_diagram(reference_json)

            # Process the object diagram with the domain model
            object_model = process_object_diagram(json_data, buml_model)

            # Generate a single file with both domain model and object model
            output_file_path = os.path.join(temp_dir, "complete_model.py")
            domain_model_to_code(model=buml_model, file_path=output_file_path, objectmodel=object_model)
            file_content = await _read_file(output_file_path, "rb")
            return Response(
                content=file_content,
                media_type="text/plain",
                headers={"Content-Disposition": 'attachment; filename="complete_model.py"'},
            )

        elif elements_data.get("type") == "AgentDiagram":
            agent_model = process_agent_diagram(json_data)
            output_file_path = os.path.join(temp_dir, "agent_buml.py")
            agent_model_to_code(agent_model, output_file_path)
            file_content = await _read_file(output_file_path, "rb")
            return Response(
                content=file_content,
                media_type="text/plain",
                headers={
                    "Content-Disposition": 'attachment; filename="agent_buml.py"'
                },
            )

        else:
            raise ValueError(
                f"Unsupported or missing diagram type: {elements_data.get('type')}"
            )


@router.post("/get-project-json-model", response_model=ProjectExportResponse)
@handle_endpoint_errors("get_project_json_model")
async def get_project_json_model(buml_file: UploadFile = File(...)):
    content = await buml_file.read()
    _validate_upload(buml_file, max_size=MAX_BUML_SIZE, allowed_extensions=ALLOWED_BUML_EXTENSIONS, content=content)
    _validate_file_content(content, buml_file.filename or "")
    buml_content = content.decode("utf-8")

    parsed_project = project_to_json(buml_content)

    return {
        "project": parsed_project,
        "exportedAt": datetime.now(timezone.utc).isoformat(),
        "version": API_VERSION
    }

@router.post("/get-json-model", response_model=DiagramExportResponse)
@handle_endpoint_errors("get_single_json_model")
async def get_single_json_model(buml_file: UploadFile = File(...)):
    """
    Convert a BUML Python file to JSON format for single diagram import.
    This endpoint handles both full project files and single diagram files.
    """
    content = await buml_file.read()
    _validate_upload(buml_file, max_size=MAX_BUML_SIZE, allowed_extensions=ALLOWED_BUML_EXTENSIONS, content=content)
    _validate_file_content(content, buml_file.filename or "")
    buml_content = content.decode("utf-8")

    # First, try to detect what type of file this is by analyzing the content
    diagram_data = None
    diagram_type = None
    safe_filename = os.path.basename(buml_file.filename) if buml_file.filename else ""
    diagram_title = safe_filename.rsplit('.', 1)[0] if safe_filename else "Imported Diagram"

    # Analyze content to determine the type
    content_lower = buml_content.lower()

    # Check for specific BUML patterns to determine the diagram type
    is_agent = any(keyword in content_lower for keyword in [
        'agent(', '.new_intent(', '.new_state(', 'when_intent_matched', 'session.reply'
    ])

    is_state_machine = any(keyword in content_lower for keyword in [
        'statemachine(', 'when_event_go_to', '.add_transition', '.new_body'
    ]) and not is_agent

    is_domain_model = any(keyword in content_lower for keyword in [
        'domainmodel(', '.create_class(', '.add_attribute(', '.create_association'
    ])

    is_gui_model = any(keyword in content_lower for keyword in [
        'guimodel(', '.new_screen(', '.new_module(', 'viewcomponent', 'viewcontainer'
    ])

    is_project = 'project(' in content_lower or 'def create_project' in content_lower

    # Try to parse based on detected type
    if is_project:
        try:
            parsed_project = project_to_json(buml_content)

            # Find the first available diagram in the project
            if parsed_project.get("ClassDiagram") and parsed_project["ClassDiagram"].get("model"):
                diagram_data = parsed_project["ClassDiagram"]
                diagram_type = "ClassDiagram"
            elif parsed_project.get("ObjectDiagram") and parsed_project["ObjectDiagram"].get("model"):
                diagram_data = parsed_project["ObjectDiagram"]
                diagram_type = "ObjectDiagram"
            elif parsed_project.get("StateMachineDiagram") and parsed_project["StateMachineDiagram"].get("model"):
                diagram_data = parsed_project["StateMachineDiagram"]
                diagram_type = "StateMachineDiagram"
            elif parsed_project.get("AgentDiagram") and parsed_project["AgentDiagram"].get("model"):
                diagram_data = parsed_project["AgentDiagram"]
                diagram_type = "AgentDiagram"
            elif parsed_project.get("GUINoCodeDiagram") and parsed_project["GUINoCodeDiagram"].get("model"):
                diagram_data = parsed_project["GUINoCodeDiagram"]
                diagram_type = "GUINoCodeDiagram"

            if diagram_data and diagram_data.get("title"):
                diagram_title = diagram_data["title"]

        except Exception as project_error:
            logger.error("Project parsing failed: %s", str(project_error))

    elif is_agent:
        try:
            logger.info("Detected Agent diagram, parsing...")
            agent_json = agent_buml_to_json(buml_content)
            diagram_data = {
                "title": diagram_title,
                "model": agent_json
            }
            diagram_type = "AgentDiagram"
        except Exception as agent_error:
            logger.error("Agent diagram parsing failed: %s", str(agent_error))

    elif is_state_machine:
        try:
            logger.info("Detected State Machine diagram, parsing...")
            state_machine_json = state_machine_to_json(buml_content)
            diagram_data = {
                "title": diagram_title,
                "model": state_machine_json
            }
            diagram_type = "StateMachineDiagram"
        except Exception as sm_error:
            logger.error("State machine parsing failed: %s", str(sm_error))

    elif is_domain_model:
        try:
            logger.info("Detected Domain Model (Class diagram), parsing...")
            domain_model = parse_buml_content(buml_content)
            if domain_model and len(domain_model.types) > 0:
                diagram_json = class_buml_to_json(domain_model)
                diagram_data = {
                    "title": diagram_title,
                    "model": diagram_json
                }
                diagram_type = "ClassDiagram"
            else:
                raise ValueError("No types found in domain model")
        except Exception as class_error:
            logger.error("Class diagram parsing failed: %s", str(class_error))

    elif is_gui_model:
        try:
            logger.info("Detected GUI Model diagram, parsing...")
            gui_json = gui_buml_to_json(buml_content)
            diagram_data = {
                "title": diagram_title,
                "model": gui_json
            }
            diagram_type = "GUINoCodeDiagram"
        except Exception as gui_error:
            logger.error("GUI diagram parsing failed: %s", str(gui_error))

    # Check if we successfully parsed any diagram
    if diagram_data is None or diagram_type is None:
        raise ValueError(
            "Could not parse BUML file. The file format was not recognized as a valid BUML diagram or project. "
            "Supported formats: ClassDiagram, ObjectDiagram, StateMachineDiagram, AgentDiagram, GUINoCodeDiagram, or Project."
        )

    # Return the diagram in the format expected by the frontend
    return {
        "title": diagram_title,
        "model": {
            **diagram_data.get("model", {}),
            "type": diagram_type
        },
        "diagramType": diagram_type,
        "exportedAt": datetime.now(timezone.utc).isoformat(),
        "version": API_VERSION
    }

@router.post("/csv-to-domain-model", response_model=DiagramExportResponse)
@handle_endpoint_errors("csv_to_domain_model_endpoint")
async def csv_to_domain_model_endpoint(files: list[UploadFile] = File(...)):
    """
    Accepts one or more CSV or XLSX files and returns a B-UML domain model as JSON.
    XLSX files are converted to CSV (first worksheet) before reverse engineering.
    """
    file_paths = []
    with tempfile.TemporaryDirectory(prefix=CSV_TEMP_DIR_PREFIX) as temp_dir:
        # Save uploaded files to temp dir
        for file in files:
            file_content = await file.read()
            _validate_upload(
                file,
                max_size=MAX_CSV_SIZE,
                allowed_extensions=ALLOWED_SPREADSHEET_EXTENSIONS,
                content=file_content,
            )
            _validate_file_content(file_content, file.filename or "")
            safe_filename = os.path.basename(file.filename)
            _, ext = os.path.splitext(safe_filename)
            ext = ext.lower()

            if ext == ".xlsx":
                csv_name = os.path.splitext(safe_filename)[0] + ".csv"
                file_path = os.path.join(temp_dir, csv_name)
                _xlsx_bytes_to_csv_file(file_content, file_path)
            else:
                file_path = os.path.join(temp_dir, safe_filename)
                await _write_file(file_path, file_content)
            file_paths.append(file_path)

        # Generate DomainModel from CSVs
        domain_model = csv_to_domain_model(file_paths, model_name="ClassDiagram")

        # Parse domain model to JSON using the same logic as get_single_json_model
        diagram_title = "ClassDiagram"
        diagram_type = "ClassDiagram"
        if domain_model and hasattr(domain_model, "types") and len(domain_model.types) > 0:
            diagram_json = class_buml_to_json(domain_model)
            diagram_data = {
                "title": diagram_title,
                "model": diagram_json
            }
        else:
            raise ValueError("No types found in domain model")

        # Check if we successfully parsed any diagram
        if diagram_data is None or diagram_type is None:
            raise ValueError(
                "Could not parse BUML file. The file format was not recognized as a valid BUML diagram or project. "
                "Supported formats: ClassDiagram, ObjectDiagram, StateMachineDiagram, AgentDiagram, GUINoCodeDiagram, or Project."
            )

        # Return the diagram in the format expected by the frontend
        return {
            "title": diagram_title,
            "model": {
                **diagram_data.get("model", {}),
                "type": diagram_type
            },
            "diagramType": diagram_type,
            "exportedAt": datetime.now(timezone.utc).isoformat(),
            "version": API_VERSION
        }

@router.post("/get-json-model-from-image", response_model=DiagramExportResponse)
@handle_endpoint_errors("get_json_model_from_image")
async def get_json_model_from_image(
    image_file: UploadFile = File(...),
    api_key: str = Form(...),
    existing_model: str = Form(None),
):
    """
    Accepts a PNG or JPEG image and an OpenAI API key, uses BESSER's imagetouml feature to transform the image into a BUML class diagram, then converts the BUML to JSON and returns the JSON object.

    When `existing_model` is provided (as a stringified diagram JSON of an existing
    ClassDiagram), the image is merged into that model — same-name classes are
    preserved, new ones are added — and the merged diagram is returned.
    """
    # Save uploaded image to a temp folder
    image_content = await image_file.read()
    _validate_upload(image_file, max_size=MAX_IMAGE_SIZE, allowed_extensions=ALLOWED_IMAGE_EXTENSIONS, content=image_content)
    _validate_file_content(image_content, image_file.filename or "")

    existing_domain_model = None
    if existing_model:
        try:
            existing_json = json.loads(existing_model)
            if existing_json.get("model", {}).get("elements") or existing_json.get("elements"):
                payload = existing_json if "model" in existing_json else {"title": "Existing", "model": existing_json}
                existing_domain_model = process_class_diagram(payload)
        except (json.JSONDecodeError, ValueError):
            existing_domain_model = None

    with tempfile.TemporaryDirectory() as temp_dir:
        safe_filename = os.path.basename(image_file.filename)
        image_path = os.path.join(temp_dir, safe_filename)
        await _write_file(image_path, image_content)

        # Create a folder for the image as expected by mockup_to_buml
        image_folder = os.path.join(temp_dir, "images")
        os.makedirs(image_folder, exist_ok=True)
        os.rename(image_path, os.path.join(image_folder, safe_filename))

        # Use image_to_buml to generate DomainModel from image and API key
        # Find the image file path
        image_files = [f for f in os.listdir(image_folder) if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
        if not image_files:
            raise HTTPException(status_code=400, detail="No valid image file found.")
        image_path = os.path.join(image_folder, image_files[0])

        domain_model = image_to_buml(
            image_path=image_path,
            openai_token=api_key,
            existing_model=existing_domain_model,
        )
        diagram_json = class_buml_to_json(domain_model)

        diagram_title = diagram_json.get("title", "Imported Class Diagram")
        diagram_type = "ClassDiagram"
        return {
            "title": diagram_title,
            "model": {**diagram_json, "type": diagram_type},
            "diagramType": diagram_type,
            "exportedAt": datetime.now(timezone.utc).isoformat(),
            "version": API_VERSION,
        }

@router.post("/get-json-model-from-kg", response_model=DiagramExportResponse)
@handle_endpoint_errors("get_json_model_from_kg")
async def get_json_model_from_kg(
    kg_file: UploadFile = File(...),
    api_key: str = Form(...)
):
    """
    Accepts an .TTL, .RDF, or .JSON knowledge graphs and an OpenAI API key, uses BESSER's kgtouml feature to transform the kg into a BUML class diagram in JSON.
    """
    # Save uploaded KG to a temp folder
    kg_content = await kg_file.read()
    _validate_upload(kg_file, max_size=MAX_BUML_SIZE, allowed_extensions=ALLOWED_KG_EXTENSIONS, content=kg_content)

    with tempfile.TemporaryDirectory() as temp_dir:
        safe_filename = os.path.basename(kg_file.filename)
        kg_path = os.path.join(temp_dir, safe_filename)
        await _write_file(kg_path, kg_content)
        # Create a folder for the kg as expected by mockup_to_buml
        kg_folder = os.path.join(temp_dir, "kgs")
        os.makedirs(kg_folder, exist_ok=True)
        os.rename(kg_path, os.path.join(kg_folder, safe_filename))
        # Use kg_to_buml to generate DomainModel from kg and API key
        # Find the kg file path
        kg_files = [f for f in os.listdir(kg_folder) if f.lower().endswith(('.ttl', '.json', '.rdf'))]
        if not kg_files:
            raise HTTPException(status_code=400, detail="No valid KG file found.")
        kg_path = os.path.join(kg_folder, kg_files[0])
        domain_model = kg_to_buml(kg_path=kg_path, openai_token=api_key)
        diagram_json = class_buml_to_json(domain_model)

        diagram_title = diagram_json.get("title", "Imported Class Diagram")
        diagram_type = "ClassDiagram"
        return {
            "title": diagram_title,
            "model": {**diagram_json, "type": diagram_type},
            "diagramType": diagram_type,
            "exportedAt": datetime.now(timezone.utc).isoformat(),
            "version": API_VERSION,
        }


@router.post("/import-owl", response_model=DiagramExportResponse)
@handle_endpoint_errors("import_owl")
async def import_owl(owl_file: UploadFile = File(...)):
    """Parse an uploaded OWL/RDF ontology into a KnowledgeGraph diagram JSON.

    Accepts ``.owl``, ``.ttl``, ``.rdf``, ``.xml``, ``.nt``, ``.n3``. Returns
    the JSON shape rendered by the KG editor's Cytoscape canvas (``nodes`` +
    ``edges``) wrapped in the standard diagram-export envelope.
    """
    owl_content = await owl_file.read()
    _validate_upload(
        owl_file,
        max_size=MAX_OWL_SIZE,
        allowed_extensions=ALLOWED_OWL_EXTENSIONS,
        content=owl_content,
    )
    _validate_file_content(owl_content, owl_file.filename or "")

    with tempfile.TemporaryDirectory(prefix=TEMP_DIR_PREFIX) as temp_dir:
        safe_filename = os.path.basename(owl_file.filename or "ontology.owl")
        owl_path = os.path.join(temp_dir, safe_filename)
        await _write_file(owl_path, owl_content)
        kg = owl_file_to_knowledge_graph(owl_path)
        diagram_json = kg_to_json(kg)

        diagram_title = diagram_json.get("title", "Imported Knowledge Graph")
        diagram_type = "KnowledgeGraphDiagram"
        return {
            "title": diagram_title,
            "model": {**diagram_json["model"], "type": diagram_type},
            "diagramType": diagram_type,
            "exportedAt": datetime.now(timezone.utc).isoformat(),
            "version": API_VERSION,
        }


def _kg_payload_to_kg(input_data: DiagramInput):
    """Validate that ``input_data`` carries a KG diagram and parse it.

    Raises:
        HTTPException(400): when the payload isn't a ``KnowledgeGraphDiagram``.
    """
    json_data = input_data.model_dump()
    model = json_data.get("model") or {}
    if model.get("type") != "KnowledgeGraphDiagram":
        raise HTTPException(
            status_code=400,
            detail="Expected a KnowledgeGraphDiagram payload (model.type must be 'KnowledgeGraphDiagram').",
        )
    return process_kg_diagram(json_data), json_data


@router.post("/kg-to-class-diagram", response_model=DiagramExportResponse)
@handle_endpoint_errors("kg_to_class_diagram")
async def kg_to_class_diagram_endpoint(input_data: DiagramInput):
    """Transform a Knowledge Graph diagram into a BUML Class Diagram (TBox).

    Deterministic, offline (no LLM) conversion: KGClass → Class,
    KGProperty (rdfs:domain/range driven) → attribute or BinaryAssociation,
    rdfs:subClassOf → Generalization. Multi-valued literal usage in the
    ABox bumps the corresponding attribute multiplicity to 0..*.
    """
    kg, _json_data = _kg_payload_to_kg(input_data)
    base_title = (input_data.title or kg.name or "Knowledge Graph")
    result = kg_to_class_diagram_buml(kg, model_name=base_title)
    diagram_json = class_buml_to_json(result.domain_model)
    return {
        "title": f"{base_title} (Class Diagram)",
        "model": {**diagram_json, "type": "ClassDiagram"},
        "diagramType": "ClassDiagram",
        "warnings": [w.__dict__ for w in result.warnings],
        "exportedAt": datetime.now(timezone.utc).isoformat(),
        "version": API_VERSION,
    }


@router.post("/kg-to-object-diagram", response_model=DiagramExportResponse)
@handle_endpoint_errors("kg_to_object_diagram")
async def kg_to_object_diagram_endpoint(input_data: DiagramInput):
    """Transform a Knowledge Graph diagram into a BUML Object Diagram (ABox).

    KGIndividual → Object (typed via rdf:type), literal-target edges → slots,
    individual-target edges → links. Blank nodes are skipped. The freshly
    derived class diagram is embedded as ``referenceDiagramData`` so the
    frontend can resolve ``classId`` references.
    """
    kg, _json_data = _kg_payload_to_kg(input_data)
    base_title = (input_data.title or kg.name or "Knowledge Graph")

    class_result = kg_to_class_diagram_buml(kg, model_name=base_title)
    object_result = kg_to_object_diagram_buml(kg, class_result=class_result, model_name=base_title)
    domain_json = class_buml_to_json(class_result.domain_model)
    domain_json = {**domain_json, "type": "ClassDiagram"}
    diagram_json = object_model_to_json(object_result.object_model, domain_json)
    return {
        "title": f"{base_title} (Object Diagram)",
        "model": diagram_json,
        "diagramType": "ObjectDiagram",
        "warnings": [w.__dict__ for w in object_result.warnings],
        "exportedAt": datetime.now(timezone.utc).isoformat(),
        "version": API_VERSION,
    }


@router.post("/transform-agent-model-json")
@handle_endpoint_errors("transform_agent_model_json")
async def transform_agent_model_json(input_data: DiagramInput):
    """Transform an agent JSON model using the generator and return personalized_agent_model as JSON."""
    with tempfile.TemporaryDirectory(prefix=f"{AGENT_TEMP_DIR_PREFIX}{uuid.uuid4().hex}_") as temp_dir:
        json_data = input_data.model_dump()
        raw_config = json_data.get("config") if isinstance(json_data.get("config"), dict) else None
        fallback_config = input_data.config if isinstance(input_data.config, dict) else None
        base_config = raw_config if raw_config is not None else fallback_config
        if base_config is None:
            raise HTTPException(status_code=400, detail="Config is required for transformation")

        logger.debug("[Agent transform] config: %s", json.dumps(base_config, indent=2, default=str))
        config = deepcopy(base_config)
        user_profile_payload = config.get("userProfileModel") if isinstance(config, dict) else None
        if isinstance(user_profile_payload, dict):
            config["userProfileModel"] = _generate_user_profile_document(user_profile_payload)
        elif isinstance(config, dict) and "userProfileModel" in config:
            config.pop("userProfileModel", None)

        json_data["config"] = config
        # Convert to BUML model
        agent_model = process_agent_diagram(json_data)

        # Persist BUML to file for generator import
        agent_file = os.path.join(temp_dir, AGENT_MODEL_FILENAME)
        agent_model_to_code(agent_model, agent_file)

        # Import the generated module (spec_from_file_location uses absolute path, no sys.path needed)
        spec = importlib.util.spec_from_file_location("agent_model", agent_file)
        agent_module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(agent_module)

        generator_info = get_generator_info("agent")
        generator_class = generator_info.generator_class
        generator_instance = generator_class(
            getattr(agent_module, "agent", agent_model),
            config=config,
            openai_api_key=extract_openai_api_key(config),
            output_dir=temp_dir,
        )
        generator_instance.generate()

        personalized_json_path = os.path.join(temp_dir, OUTPUT_DIR_NAME, "personalized_agent_model.json")
        if not os.path.isfile(personalized_json_path):
            raise HTTPException(status_code=500, detail="personalized_agent_model.json not found after generation")

        personalized_json = await _read_json_file(personalized_json_path)

        return {
            "model": personalized_json,
            "diagramType": "AgentDiagram",
            "exportedAt": datetime.now(timezone.utc).isoformat(),
            "version": API_VERSION,
        }


def _generate_user_profile_document(user_profile_model: dict) -> dict:
    """Generate the normalized JSON document for a stored user profile diagram.

    This is a local helper used by transform_agent_model_json. It delegates to
    the generation router's implementation for the actual user profile generation.
    """
    # Import here to avoid circular imports at module level
    from besser.utilities.web_modeling_editor.backend.routers.generation_router import (
        _generate_user_profile_document as _gen_user_profile_doc,
    )
    return _gen_user_profile_doc(user_profile_model)
